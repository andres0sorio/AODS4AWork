from openpyxl import load_workbook
import pandas as pd
import json


class ExcelReader:
    """

    """
    def __init__(self, sheet):
        """

        :param sheet:
        """
        self.sheet = sheet

        f = open("input/files_description.json", "r")
        files_description = json.loads(f.read())

        for description in files_description:
            if description["sheet"] == sheet:
                self.columns = description["columns"]
                self.data = description["data"]
                self.id = description["id"]

    def look_for_sheet(self, input_file):

        wb = load_workbook(input_file)
        sheet_names = wb.sheetnames
        if self.sheet in sheet_names:
            return True
        else:
            return False

    def combine_sheets(self, file_list):
        df = pd.DataFrame()

        f = open(file_list, "r")
        data = json.loads(f.read())

        try:
            for dt in data:
                if dt[self.sheet] == 1:
                    file = dt["file_path"] + dt["file_name"]
                    print(file)
                    df = df.append(pd.read_excel(file, sheet_name=self.sheet), ignore_index=True)

            print(len(df))
            df = df.drop_duplicates(subset=[self.id])
            df = df.sort_values(by=self.id)
            print(len(df))

            print(df.info())

            df.to_csv("output/output.csv", sep=';', encoding='utf-8-sig', index=False)

        except Exception as error:
            print(error)

